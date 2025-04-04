## Affiliation: SW-Lab, Dept of CFS, NTNU
## Date: 31.03.2025
## Author: MY Sia (with lots of help from the web, see README.md for more)

##import necessary packages
import glob
from pathlib import Path
import os ##to rename files, where necessary
from moviepy.editor import * ##MoviePy v1.0.3
from datetime import datetime ##to calculate time difference between videos
import cv2 ##so that videos can be resized properly
from openpyxl import load_workbook ##if we pass arguments from an excel file

def merge(folder, children, camera):
    """
    FUNCTION 1: CONCATENATE short videos into a long one
    Date: First written on 23.06.2024, further editing on 12.02.2025
    Input: Multiple short videos stored in subfolders.
    Output: A single video per camera per child.
    Required directory: project folder -> 'child' subfolder -> 'camera' subfolder -> short videos
    """
    for child in children:
        for cam in camera:
            video_list = []
            files_path = Path(f"{folder}/{child}/{cam}")
            file_list = glob.glob(f"{files_path}/*.mp4")
            ##check for nonexistent files
            if len(file_list) < 2:
                print(f"Nothing to join in {child}'s {cam} folder.")
            ##check for recordings that passed the hour (e.g., 10:59 to 11:00)
            elif  file_list[0][-22:-20] == "\\0" and file_list[-1][-22:-20] == "\\5":
                for f_name in file_list:
                    p1 = f_name[0:-22]
                    xx = f_name[-22:-20]
                    p2 = f_name[-20:]
                    ##rename 0 to 6 and 1 to 7
                    if xx == "\\0":
                        new_fname = f"{p1}\\6{p2}"
                        os.rename(f_name, new_fname)
                    elif xx == "\\1":
                        new_fname = f"{p1}\\7{p2}"
                        os.rename(f_name, new_fname)
                camera.append(cam)
            ##load & merge videos
            else:
                for i in file_list:
                    clip = VideoFileClip(i)
                    video_list.append(clip)
                joined = concatenate_videoclips(video_list)
                joined = joined.set_fps(fps=30)
                ##render output
                first_frame = file_list[0][-21:-15]
                cam = cam.lower()
                joined.write_videofile(f"{folder}/{child}/{child}_{cam}_{first_frame}.mp4")


def overlay(folder, attempts, bgcam, topcam, newname, propsize, dur,
            excel, children, start, end, corr):
    """
    FUNCTION 2: SYNC AND OVERLAY a downsized 'top' video on a 'base' video
    Date: First written on 23.06.2024
    Input: Two videos, one showing the participant and another showing what the participant saw.
    Output: A composite video with one video as the base and another overlaid on the top left corner.
    Required directory: project folder -> 'child' subfolder -> videos
    """
    ##extract information if we load an excel file
    if excel != None:
        wb = load_workbook(excel)
        sheet = wb["Sheet1"]

        children=[] ##which child subfolder are we processing
        list_children = sheet["a"]
        for i in list_children[1:]:
            children.append(i.value)

        start=[] ##the seconds at which the task STARTED
        list_start =  sheet["b"]
        for i in list_start[1:]:
            start.append(i.value)

        end=[] ##the seconds at which the task ENDED
        list_end =  sheet["c"]
        for i in list_end[1:]:
            end.append(i.value)

        corr=[] ##manually correct out-of-sync videos
        list_corr =  sheet["d"]
        for i in list_corr[1:]:
            corr.append(i.value)

    ##if the task duration is the same for all participants, use it to calculate the end time
    if dur != None:
        end = []
        for i in start:
            end.append(i + dur)

    ##prepare to overlay videos
    n = 0
    for child in children:
        vid_path = Path(f"{folder}/{child}/")
        ##check for the base video
        baby_list = glob.glob(f"{vid_path}/*{bgcam}*.mp4")
        if len(baby_list) < 1:
            print(f"Can't find {child}'s {bgcam}. Please check the path and folder names.")
        else:
            baby_vid = VideoFileClip(baby_list[0])
            t_baby = baby_list[0][-10:-5]
            t_baby = t_baby.replace("M", ":")
            t_baby = datetime.strptime(t_baby, "%M:%S")
            ##check for the top video
            screen_list = glob.glob(f"{vid_path}/*{topcam}*.mp4")
            if len(screen_list) < 1:
                print(f"Can't find {child}'s {topcam}. Please check the path and folder names.")
            else:
                screen_vid = VideoFileClip(screen_list[0])
                screen_vid = screen_vid.resize(propsize).margin(5) ##add a 5px border
                t_screen = screen_list[0][-10:-5]
                t_screen = t_screen.replace("M", ":")
                t_screen = datetime.strptime(t_screen, "%M:%S")
                ##calculate the time difference between videos
                diff = t_screen - t_baby
                diff = diff.total_seconds()
                ##sync & overlay videos
                if attempts == 1: ##first round
                    x = 0 ##no manual correction
                    all_vid = CompositeVideoClip([baby_vid.subclip(start[n]+diff, end[n]+diff),
                                                screen_vid.subclip(start[n], end[n]).set_position((0, 50))])
                elif attempts > 1: ##corrective round
                    x = corr[n]
                    all_vid = CompositeVideoClip([baby_vid.subclip(start[n]+diff+corr[n], end[n]+diff+corr[n]),
                                                screen_vid.subclip(start[n], end[n]).set_position((0, 50))])
                ##render output
                all_vid.write_videofile(f"{folder}/{child}/{child}_{newname}_{attempts}_corr={x}.mp4")
                n += 1 ##now, do the next one


def crop(folder, cam, newname, dur, amplify,
         excel, children, start, end, x1, x2, y1, y2):
    """
    FUNCTION 3: CROP a video
    Date: First written on 18.02.2025
    Input: A video.
    Output: A cropped (and possibly clipped and amplified) video.
    Required directory: project folder -> 'child' subfolder -> videos
    """
    ##extract information if we load an excel file
    if excel != None:
        wb = load_workbook(excel)
        sheet = wb["Sheet1"]

        children=[] ##which child subfolder are we processing
        list_children = sheet["a"]
        for i in list_children[1:]:
            children.append(i.value)

        start=[] ##the seconds at which the task STARTED
        list_start =  sheet["b"]
        for i in list_start[1:]:
            start.append(i.value)

        end=[] ##the seconds at which the task ENDED
        list_end =  sheet["c"]
        for i in list_end[1:]:
            end.append(i.value)

        x1=[] ##the dimension of the cropped area
        list_x1 =  sheet["d"]
        for i in list_x1[1:]:
            x1.append(i.value)

        x2=[]
        list_x2 =  sheet["e"]
        for i in list_x2[1:]:
            x2.append(i.value)

        y1=[]
        list_y1 =  sheet["f"]
        for i in list_y1[1:]:
            y1.append(i.value)

        y2=[]
        list_y2 =  sheet["g"]
        for i in list_y2[1:]:
            y2.append(i.value)

    ##if the task duration is the same for all participants, use it to calculate the end time
    if dur != None:
        end = []
        for i in start:
            end.append(i + dur)

    ##prepare to crop videos
    n = 0
    for child in children:
        vid_path = Path(f"{folder}/{child}/")
        vid_list = glob.glob(f"{vid_path}/*{cam}*.mp4")
        ##check for nonexistent files
        if len(vid_list) < 1: 
            print(f"Can't find {child}'s {cam}. Please check the path and folder names.")
        ##load, clip, crop, & amplify videos
        else:
            thevid = VideoFileClip(vid_list[0])
            if thevid.duration > end[n]:
                thevid = thevid.subclip(start[n], end[n])
            else:
                thevid = thevid.subclip(start[n], thevid.duration)
            cropped = thevid.crop(x1=x1[n], x2=x2[n], y1=y1[n], y2=y2[n])
            cropped = cropped.volumex(amplify)
            ##render output
            cropped.write_videofile(f"{folder}/{child}/{child}_{newname}.mp4")
            n += 1 ##now, do the next one


def join2side(folder, attempts, cam1, cam2, newname, dur, amplify_who, amplify, mute_who, crop_who,
              excel, children, main, start, end, corr, x1, x2, y1, y2):
    """
    FUNCTION 4A: JUXTAPOSE two videos
    Date: First written on 08.07.2024, further editing on 18.02.2025
    Input: Two video recordings of a participant performing a task (from different angles).
    Output: A single video with the two recordings beside each other (one recording is set larger than the other).
    Required directory: project folder -> 'child' subfolder -> videos
    """
    ##extract information if we load an excel file
    if excel != None:
        wb = load_workbook(excel)
        sheet = wb["Sheet1"]

        children=[] ##which child subfolder are we processing
        list_children = sheet["a"]
        for i in list_children[1:]:
            children.append(i.value)

        main=[] ##the video with the best angle
        list_main = sheet["b"]
        for i in list_main[1:]:
            main.append(i.value)

        start=[] ##the seconds at which the task STARTED
        list_start = sheet["c"]
        for i in list_start[1:]:
            start.append(i.value)

        end=[] ##the seconds at which the task ENDED
        list_end = sheet["d"]
        for i in list_end[1:]:
            end.append(i.value)

        corr=[] ##manually correct out-of-sync videos
        list_corr = sheet["e"]
        for i in list_corr[1:]:
            corr.append(i.value)

        x1=[] ##the dimension of the cropped area
        list_x1 = sheet["f"]
        for i in list_x1[1:]:
            x1.append(i.value)

        x2=[]
        list_x2 = sheet["g"]
        for i in list_x2[1:]:
            x2.append(i.value)

        y1=[]
        list_y1 = sheet["h"]
        for i in list_y1[1:]:
            y1.append(i.value)

        y2=[]
        list_y2 = sheet["i"]
        for i in list_y2[1:]:
            y2.append(i.value)

    ##if the task duration is the same for all participants, use it to calculate the end time
    if dur != None:
        end = []
        for i in start:
            end.append(i + dur)

    ##start juxtaposing videos
    n = 0
    for child in children:
        vid_path = Path(f"{folder}/{child}/")
        ##check for the first camera
        vid1_list = glob.glob(f"{vid_path}/*{cam1}*.mp4")
        if len(vid1_list) < 1:
            print(f"Can't find {child}'s {cam1}. Please check the path and folder names.")
        else:
            vid1 = VideoFileClip(vid1_list[0])
            if crop_who == cam1:
                vid1 = vid1.crop(x1=x1[n], x2=x2[n], y1=y1[n], y2=y2[n])
            ##check for the second camera
            vid2_list = glob.glob(f"{vid_path}/*{cam2}*.mp4")
            if len(vid2_list) < 1:
                print(f"Can't find {child}'s {cam2}. Please check the path and folder names.")
            else:
                vid2 = VideoFileClip(vid2_list[0])
                if crop_who == cam2:
                    vid2 = vid2.crop(x1=x1[n], x2=x2[n], y1=y1[n], y2=y2[n])
                ##calculate the time difference between videos
                t_vid1 = vid1_list[0][-10:-5]
                t_vid1 = t_vid1.replace("M", ":")
                t_vid1 = datetime.strptime(t_vid1, "%M:%S")
                t_vid2 = vid2_list[0][-10:-5]
                t_vid2 = t_vid2.replace("M", ":")
                t_vid2 = datetime.strptime(t_vid2, "%M:%S")
                diff = t_vid1 - t_vid2
                diff = diff.total_seconds()
                ##sync & clip videos
                if vid1.duration > end[n]:
                    vid1 = vid1.subclip(start[n], end[n])
                else:
                    vid1 = vid1.subclip(start[n], vid1.duration)
                if attempts == 1: ##first round
                    x = 0 ##no manual correction
                    if vid2.duration > end[n]+diff:
                        vid2 = vid2.subclip(start[n]+diff, end[n]+diff)
                    else:
                        vid2 = vid2.subclip(start[n]+diff, vid2.duration)
                    ##render audio files to aid correction of out-of-sync videos
                    audio_vid1 = vid1.audio
                    audio_vid1.write_audiofile(f"{folder}/{child}/{child}_{cam1}_audio.mp3")
                    audio_vid2 = vid2.audio
                    audio_vid2.write_audiofile(f"{folder}/{child}/{child}_{cam2}_audio.mp3")
                elif attempts > 1: ##corrective round
                    x = corr[n]
                    if vid2.duration > end[n]+diff+corr[n]:
                        vid2 = vid2.subclip(start[n]+diff+corr[n], end[n]+diff+corr[n])
                    else:
                        vid2 = vid2.subclip(start[n]+diff+corr[n], vid2.duration)
                ##amplify (or mute) & add a blue border
                if amplify_who == cam1:
                    vid1 = vid1.volumex(amplify).margin(10, color=(0, 0, 225))
                    vid2 = vid2.margin(10)
                elif amplify_who == cam2:
                    vid2 = vid2.volumex(amplify).margin(10, color=(0, 0, 225))
                    vid1 = vid1.margin(10)
                if mute_who == cam1:
                    vid1 = vid1.volumex(0)
                elif mute_who == cam2:
                    vid2 = vid2.volumex(0)
                ##resize videos
                if main[n] == cam1:
                    major, minor = vid1, vid2
                else:
                    major, minor = vid2, vid1
                major_vid = major.resize(0.7)
                minor_vid = minor.resize(0.4)
                final_vid = clips_array([[major_vid, minor_vid], ])
                ##render output
                final_vid.write_videofile(f"{folder}/{child}/{child}_{newname}_{attempts}_corr={x}.mp4")
                n += 1 ##now, do the next one


def join3side():
    """
    FUNCTION 4B: JUXTAPOSE three videos
    Date: first written on 08.07.2024
    Input: Three video recordings of a participant performing a task (from different angles).
    Output: A single video with the three recordings beside each other (one recording is set larger than the other two).
    Required directory: project folder -> 'child' subfolder -> videos
    """
    print("coming soon")

